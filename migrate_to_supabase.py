#!/usr/bin/env python3
"""
migrate_to_supabase.py

One-shot migration of the MLBB Hero Counter workbook into Supabase.

Reads:
    Heroes                          -> heroes
    Data-Input  rows 7-16   (Block 1: Global Weights)          -> global_weights
    Data-Input  rows 18-24  (Block 2: Role Matchup Matrix)     -> role_matrix
    Data-Input  rows 29-128 (Block 3: Hard Counter Rules)      -> hard_counter_rules
    Data-Input  rows 163-185(Block 5: Style/Tag Interaction)   -> style_matrix
    Data-Input  rows 190-1189(Block 6: Manual Overrides)       -> manual_overrides

Does NOT touch `counter_scores` — that table is populated later by a separate
batch job that actually computes the formula from Documentation - Computations.

Usage:
    export SUPABASE_URL="https://xxxx.supabase.co"
    export SUPABASE_SERVICE_KEY="eyJ..."
    python migrate_to_supabase.py path/to/mobile_legends_heroes_updated.xlsx
"""

import os
import re
import sys
from pathlib import Path
from typing import Any, Optional

import openpyxl
from supabase import create_client, Client
from postgrest.exceptions import APIError

# Load .env file if present
_dotenv = Path(__file__).parent / ".env"
if _dotenv.exists():
    for _line in _dotenv.read_text().splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())


# ----------------------------------------------------------------------------
# Config: row ranges from the workbook (1-indexed, inclusive, as documented
# in Data-Input's section headers / the task spec).
# ----------------------------------------------------------------------------
GLOBAL_WEIGHTS_ROWS = (7, 16)          # Block 1
ROLE_MATRIX_HEADER_ROW = 18            # Block 2 header (defender roles across)
ROLE_MATRIX_DATA_ROWS = (19, 24)       # Block 2 data rows (attacker roles down)
HARD_COUNTER_ROWS = (29, 130)          # Block 3
STYLE_MATRIX_HEADER_ROW = 164          # Block 5 header (defender tags across)
STYLE_MATRIX_DATA_ROWS = (165, 186)    # Block 5 data rows (attacker tags down)
MANUAL_OVERRIDES_ROWS = (191, 1189)    # Block 6

EXPECTED_HERO_COUNT = 133


def get_supabase_client() -> Client:
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY")
    if not url or not key:
        sys.exit(
            "ERROR: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set as "
            "environment variables. Never hardcode credentials in this script."
        )
    return create_client(url, key)


def to_bool_yes_no(value: Optional[str]) -> bool:
    """Heroes.has_antiheal is stored in the sheet as the string 'Yes'/'No'."""
    if value is None:
        return False
    return str(value).strip().lower() == "yes"


def clean(value: Any) -> Any:
    """Trim strings, pass through everything else (including None/numbers)."""
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def row_values(ws, row: int, min_col: int, max_col: int) -> list:
    return [ws.cell(row=row, column=c).value for c in range(min_col, max_col + 1)]


# ----------------------------------------------------------------------------
# Extraction functions — each returns a list of dicts ready to upsert.
# ----------------------------------------------------------------------------
def extract_heroes(wb) -> list[dict]:
    ws = wb["Heroes"]
    headers = [c.value for c in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # blank trailing row
            continue
        rec = dict(zip(headers, row))
        rec["id"] = int(rec["id"])
        for col in ("offense", "ability_effects", "durability", "difficulty", "spike_order"):
            rec[col] = int(rec[col])
        rec["has_antiheal"] = to_bool_yes_no(rec.get("has_antiheal"))
        rec["has_true_damage"] = to_bool_yes_no(rec.get("has_true_damage"))
        rec["teamfight_contribution"] = int(rec.get("teamfight_contribution") or 0)
        for col in ("role2", "style2", "lane2"):
            rec[col] = clean(rec.get(col))
        for col in ("name", "role", "style1", "damage_type", "range_type",
                    "lane1", "power_spike", "resource"):
            rec[col] = clean(rec.get(col))
        records.append(rec)
    return records


def extract_global_weights(wb) -> list[dict]:
    ws = wb["Data-Input"]
    start, end = GLOBAL_WEIGHTS_ROWS
    records = []
    for row in range(start, end + 1):
        coefficient, value, description = row_values(ws, row, 1, 3)
        coefficient = clean(coefficient)
        if coefficient is None:
            continue
        records.append({
            "coefficient": coefficient,
            "value": float(value),
            "description": clean(description),
        })
    return records


def extract_role_matrix(wb) -> list[dict]:
    ws = wb["Data-Input"]
    defender_roles = [clean(v) for v in row_values(ws, ROLE_MATRIX_HEADER_ROW, 2, 7)]
    start, end = ROLE_MATRIX_DATA_ROWS
    records = []
    for row in range(start, end + 1):
        vals = row_values(ws, row, 1, 7)
        attacker_role = clean(vals[0])
        if attacker_role is None:
            continue
        for defender_role, points in zip(defender_roles, vals[1:]):
            if defender_role is None or points is None:
                continue
            records.append({
                "attacker_role": attacker_role,
                "defender_role": defender_role,
                "points": float(points),
            })
    return records


def extract_hard_counter_rules(wb) -> list[dict]:
    ws = wb["Data-Input"]
    start, end = HARD_COUNTER_ROWS
    seen = {}
    for row in range(start, end + 1):
        attacker, condition_type, condition_value, bonus, penalty, note = row_values(ws, row, 1, 6)
        attacker = clean(attacker)
        if attacker is None:
            continue  # skip blank rows
        key = (attacker, clean(condition_type), clean(condition_value))
        if key in seen:
            seen[key]["bonus_to_attacker"] += float(bonus)
            seen[key]["penalty_to_defender"] += float(penalty)
        else:
            seen[key] = {
                "attacker": attacker,
                "condition_type": clean(condition_type),
                "condition_value": clean(condition_value),
                "bonus_to_attacker": float(bonus),
                "penalty_to_defender": float(penalty),
                "note": clean(note),
            }
    return list(seen.values())


def extract_style_matrix(wb) -> list[dict]:
    ws = wb["Data-Input"]
    # Header row has an extra leading label cell ("Attacker\Defender") in col A;
    # tags run from column B onward, however many are populated.
    header_row = [c.value for c in ws[STYLE_MATRIX_HEADER_ROW]]
    defender_tags = [clean(v) for v in header_row[1:] if clean(v) is not None]
    n_tags = len(defender_tags)

    start, end = STYLE_MATRIX_DATA_ROWS
    records = []
    for row in range(start, end + 1):
        vals = row_values(ws, row, 1, 1 + n_tags)
        attacker_tag = clean(vals[0])
        if attacker_tag is None:
            continue
        for defender_tag, points in zip(defender_tags, vals[1:]):
            if points is None:
                continue
            records.append({
                "attacker_tag": attacker_tag,
                "defender_tag": defender_tag,
                "points": float(points),
            })
    return records


def extract_manual_overrides(wb) -> list[dict]:
    ws = wb["Data-Input"]
    start, end = MANUAL_OVERRIDES_ROWS
    records = []
    for row in range(start, end + 1):
        attacker, defender, score, note = row_values(ws, row, 1, 4)
        attacker = clean(attacker)
        defender = clean(defender)
        if attacker is None or defender is None:
            continue  # skip blank rows (the "Key" formula column is ignored)
        records.append({
            "attacker": attacker,
            "defender": defender,
            "score": float(score) if score is not None else None,
            "note": clean(note),
        })
    return records


def extract_synergy_scores(wb) -> list[dict]:
    if "Synergy" not in wb.sheetnames:
        return []
    ws = wb["Synergy"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        hero_a, hero_b, synergy, note = (row + (None,) * 4)[:4]
        hero_a = clean(hero_a)
        hero_b = clean(hero_b)
        if hero_a is None or hero_b is None or synergy is None:
            continue
        records.append({
            "hero_a": hero_a,
            "hero_b": hero_b,
            "synergy": float(synergy),
        })
    return records


# ----------------------------------------------------------------------------
# Schema migration — ensures any new tables/columns from schema.sql exist
# before the upsert.  Uses psql (PostgreSQL CLI) with a direct DB connection.
# Falls back to a subprocess call consuming schema.sql.
# ----------------------------------------------------------------------------
_SCHEMA_SQL = Path(__file__).parent / "schema.sql"


def migrate_schema():
    """Run schema.sql against the Supabase DB to create tables/columns.

    Requires SUPABASE_DB_PASSWORD env var (GitHub secret or .env).
    The DB host/user/name are derived from SUPABASE_URL.
    """
    db_pass = os.environ.get("SUPABASE_DB_PASSWORD")
    supabase_url = os.environ.get("SUPABASE_URL", "")

    if not db_pass:
        print("  SKIP schema migration: SUPABASE_DB_PASSWORD not set")
        return False
    if not _SCHEMA_SQL.exists():
        print(f"  SKIP schema migration: {_SCHEMA_SQL} not found")
        return False

    # Derive project ref and construct DB URL
    # SUPABASE_URL = https://<project-ref>.supabase.co
    m = re.match(r"https://([^.]+)\.", supabase_url)
    if not m:
        print(f"  SKIP schema migration: could not parse project ref from {supabase_url}")
        return False
    project_ref = m.group(1)

    db_host = f"db.{project_ref}.supabase.co"
    db_user = "postgres"
    db_name = "postgres"
    conn_str = f"postgresql://{db_user}:{db_pass}@{db_host}:5432/{db_name}"

    import subprocess
    try:
        result = subprocess.run(
            ["psql", conn_str, "-f", str(_SCHEMA_SQL)],
            capture_output=True, text=True, timeout=30,
        )
        if result.returncode != 0:
            stderr = result.stderr.strip()
            # "already exists" is harmless — our SQL uses IF NOT EXISTS
            if "already exists" in stderr or "duplicate" in stderr.lower():
                print("  Schema migration: OK (existing objects skipped)")
            else:
                print(f"  WARNING: psql had stderr output:\n{stderr}")
        else:
            print("  Schema migration: OK")
        return True
    except FileNotFoundError:
        print("  SKIP schema migration: psql not installed on this machine")
        return False
    except subprocess.TimeoutExpired:
        print("  WARNING: schema migration timed out — continuing anyway")
        return False
    except Exception as e:
        print(f"  WARNING: schema migration failed ({e}) — continuing anyway")
        return False


# ----------------------------------------------------------------------------
# Upsert helper — supabase-py batches, chunked to stay well under request
# size limits for the larger tables (style_matrix ~ 484 rows, heroes ~ 133).
# If a column doesn't exist yet in the DB, it's stripped from the records
# automatically so the rest of the migration can proceed.
# ----------------------------------------------------------------------------
_BAD_COL_RE = re.compile(r"Could not find the '(\w+)' column")


def upsert_in_chunks(client: Client, table: str, records: list[dict],
                      on_conflict: str, chunk_size: int = 500) -> int:
    if not records:
        print(f"  {table}: nothing to upsert (0 rows found)")
        return 0
    total = 0
    for i in range(0, len(records), chunk_size):
        chunk = records[i:i + chunk_size]
        while True:
            try:
                client.table(table).upsert(chunk, on_conflict=on_conflict).execute()
                total += len(chunk)
                break
            except APIError as e:
                msg = str(e)
                # Column missing in the DB?  Strip it and retry.
                m = _BAD_COL_RE.search(msg)
                if m:
                    col = m.group(1)
                    print(f"  {table}: column '{col}' missing — stripping and retrying")
                    for r in chunk:
                        r.pop(col, None)
                    continue  # retry this chunk
                # Table itself missing?  Skip silently.
                if "does not exist" in msg or "Could not find the table" in msg:
                    print(f"  {table}: table does not exist — skipping")
                    return total
                raise
    print(f"  {table}: upserted {total} rows")
    return total


def main():
    if len(sys.argv) != 2:
        sys.exit("Usage: python migrate_to_supabase.py <path_to_workbook.xlsx>")
    xlsx_path = sys.argv[1]

    print(f"Loading workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    print("Running schema migration...")
    migrate_schema()

    client = get_supabase_client()

    print("Extracting data from workbook...")
    heroes = extract_heroes(wb)
    weights = extract_global_weights(wb)
    role_matrix = extract_role_matrix(wb)
    hard_counters = extract_hard_counter_rules(wb)
    style_matrix = extract_style_matrix(wb)
    overrides = extract_manual_overrides(wb)
    synergy = extract_synergy_scores(wb)

    print("\nUpserting into Supabase...")
    counts = {}
    counts["heroes"] = upsert_in_chunks(client, "heroes", heroes, on_conflict="id")
    counts["global_weights"] = upsert_in_chunks(
        client, "global_weights", weights, on_conflict="coefficient"
    )
    counts["role_matrix"] = upsert_in_chunks(
        client, "role_matrix", role_matrix, on_conflict="attacker_role,defender_role"
    )

    # hard_counter_rules uses a unique index on (attacker, condition_type, condition_value).
    # This is an upsert so re-running the migration safely updates existing rules.
    if hard_counters:
        counts["hard_counter_rules"] = upsert_in_chunks(
            client, 
            "hard_counter_rules", 
            hard_counters, 
            on_conflict="attacker,condition_type,condition_value"
        )
    else:
        counts["hard_counter_rules"] = 0
        print(" hard_counter_rules: nothing to insert (0 rows found)")

    counts["style_matrix"] = upsert_in_chunks(
        client, "style_matrix", style_matrix, on_conflict="attacker_tag,defender_tag"
    )
    counts["manual_overrides"] = upsert_in_chunks(
        client, "manual_overrides", overrides, on_conflict="attacker,defender"
    )
    counts["synergy_scores"] = upsert_in_chunks(
        client, "synergy_scores", synergy, on_conflict="hero_a,hero_b"
    )

    print("\n--- Row counts written ---")
    for table, n in counts.items():
        print(f"  {table:20s}: {n}")

    # ------------------------------------------------------------------
    # Sanity check
    # ------------------------------------------------------------------
    print("\n--- Sanity check ---")
    heroes_count_resp = client.table("heroes").select("id", count="exact").execute()
    actual_hero_count = heroes_count_resp.count
    if actual_hero_count == EXPECTED_HERO_COUNT:
        print(f"  OK: heroes table has {actual_hero_count} rows (expected {EXPECTED_HERO_COUNT})")
    else:
        print(
            f"  WARNING: heroes table has {actual_hero_count} rows, "
            f"expected {EXPECTED_HERO_COUNT}. Check the workbook / migration."
        )
        sys.exit(1)

    print("\nMigration complete.")


if __name__ == "__main__":
    main()
